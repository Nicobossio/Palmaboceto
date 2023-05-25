#Libreria que se usa para extraer los datos del excel
from openpyxl import load_workbook

#ruta de la carpeta del arcivo
file_path = 'Prueba.xslx'

#Acceda al archivo
wb = load_workbook(file_path)

#Accede a la hoja del excel
ws = wb['Sheet1']

#Comienza su escritura desde el punto indicado
ws['B2'] = 123

wb.save(file_path)

