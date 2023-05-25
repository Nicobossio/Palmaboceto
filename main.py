#Librerias a usar para el escaneo de los QR
##creado para los argumentos de la creacion de los excel
##se les indica via consola
#Libreria que se usa para extraer los datos del excel
import cv2
import numpy as np
import sys
from openpyxl import load_workbook

#Indicaciones para formar el excel
_, file_path, sheet, column, row = sys.argv

#Se ponen asi porque queremos que esto escriba celda por celda
row = int(row) 

#ruta de la carpeta del arcivo
file_path = 'Prueba.xlsx'

#Acceda al archivo
wb = load_workbook(file_path)

#Accede a la hoja del excel
ws = wb[sheet]


#Toma de la foto y escaneo de la imagen
captura = cv2.VideoCapture(0)

#Bucle mientras la camara esta activa
while(captura.isOpened()):
    ret, frame = captura.read()

    ##salida del programa cuando se tiene la entrada del usuario
    if (cv2.waitKey(1) == ord('s')):
        break
    qrdetection = cv2.QRCodeDetector()
    data, bbox, RectifiedImage = qrdetection.detectAndDecode(frame)#Datos que nos va a devolver el escaneo del QR


    #Si el codigo tiene un valor, se va a imprimir por pantalla
    #Tambien se valida el frame para que se escanee el QR
    if len(data) > 0:
        save = ''
        while save not in ['s', 'n']:
            save = input(f'El dato leido es: {data}, quiere guardarlo [s/n]')#da la posibilidad de poder escoger si queremos guardar o no el dato
            if save == 's':
                #Escribe el dato
                ws[f'{column}{row}'] = data
        cv2.imshow('Webcam', RectifiedImage)
    else:
        cv2.imshow('Webcam', frame)

#Se guarda el excel antes de que la camara se cierre
wb.save(file_path)


#Liberacion de la camara para el escaneo
captura.release()
cv2.destroyAllWindows   



